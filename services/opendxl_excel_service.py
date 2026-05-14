import pandas as pd
import requests
from datetime import datetime
from openpyxl import load_workbook

from utils.constants import PAYABLE_MAP


def download_template(template_file):

    download_url = template_file.get(
        "@microsoft.graph.downloadUrl"
    )

    local_path = (
        "Closing_Forms_Template.xlsx"
    )

    response = requests.get(download_url)

    with open(local_path, "wb") as f:
        f.write(response.content)

    return local_path


def populate_excel_template(
    template_path,
    output_path,
    closing_ticket_df,
    invoice_df
):

    wb = load_workbook(template_path)

    ws = wb[
        "Closing Check Transmittal Form"
    ]

    row = closing_ticket_df.iloc[0]
    
    purchase_price = row.get(
        "cr109_saleprice",
        ""
    )

    closing_date = row.get(
        "cr7de_closingdate",
        ""
    )

    seller_tcode = row.get(
        "cr7de_sellertcode",
        ""
    )

    property_address = row.get(
        "cr7de_buildingaddress",
        ""
    )

    unit = row.get(
        "cr7de_unitnumber",
        ""
    )

    seller1_name = row.get(
        "cr7de_sellername",
        ""
    )
    deal=row.get(
        "cr7de_deal",
        ""
    )
    buyer1_name = row.get(
        "cr7de_buyername",
        ""
    )
    shares=row.get(
        "cr109_shares",
        "" )
    Closing_Agent=row.get(
        "cr7de_closingagentname",
        ""
    )
    Closing_Agent_Phone=row.get(
        "cr7de_closingagentphone",
        "")
    Closing_Agent_Email=row.get(
        "cr7de_closingagentemail",
        ""    )
    Closing_Agent_Title=row.get(
        "cr7de_titlerole",
        ""    )
    Notes=row.get(
        "cr7de_notes",
        ""   )
    current_date = datetime.now().strftime(
        "%m/%d/%Y"
    )
    
    try:

        if closing_date:

            closing_date = (
                pd.to_datetime(closing_date)
                .strftime("%m/%d/%Y")
            )

    except:
        pass

    ws["D1"]=current_date
    ws["D2"] = purchase_price
    ws["D3"]= deal
    ws["D4"]=shares
    ws["D5"] = closing_date
    ws["D6"] = seller_tcode
    ws["D7"] = property_address
    ws["D8"] = unit
    ws["C13"] = seller1_name
    ws["C43"]= buyer1_name
    ws["B86"]=Closing_Agent
    ws["B87"]=Closing_Agent_Email
    ws["B88"]=Closing_Agent_Phone
    ws["B89"]=Closing_Agent_Title
    ws["B91"]=Notes
    seller_df = invoice_df[
        invoice_df["cr7de_paidby"] == 716070000
    ]

    buyer_df = invoice_df[
        invoice_df["cr7de_paidby"] == 716070001
    ]

    seller_start_row = 15

    for idx, (_, inv_row) in enumerate(
        seller_df.iterrows()
    ):

        excel_row = seller_start_row + idx

        ws[f"A{excel_row}"] = inv_row.get(
            "cr7de_chequenumber",
            ""
        )

        ws[f"B{excel_row}"] = inv_row.get(
            "cr7de_dueatclosing",
            ""
        )

        ws[f"C{excel_row}"] = inv_row.get(
            "cr7de_amount",
            ""
        )

        payable = inv_row.get(
            "cr7de_payableto",
            ""
        )

        ws[f"D{excel_row}"] = PAYABLE_MAP.get(
            payable,
            ""
        )

    buyer_start_row = 45

    for idx, (_, inv_row) in enumerate(
        buyer_df.iterrows()
    ):

        excel_row = buyer_start_row + idx

        ws[f"A{excel_row}"] = inv_row.get(
            "cr7de_chequenumber",
            ""
        )

        ws[f"B{excel_row}"] = inv_row.get(
            "cr7de_dueatclosing",
            ""
        )

        ws[f"C{excel_row}"] = inv_row.get(
            "cr7de_amount",
            ""
        )

        payable = inv_row.get(
            "cr7de_payableto",
            ""
        )

        ws[f"D{excel_row}"] = PAYABLE_MAP.get(
            payable,
            ""
        )

    wb.save(output_path)

    print("✅ Excel Populated Successfully")
